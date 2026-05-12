"""
Document Ingestion Layer — Unified parsers for PDF, TXT, and Excel files.
Extracts text and metadata from each format into a common Document structure.
"""

import os
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
import pdfplumber
import openpyxl


@dataclass
class Document:
    content: str
    metadata: dict = field(default_factory=dict)
    source: str = ""
    doc_type: str = ""


class PDFParser:
    """Extracts text from PDF files using pdfplumber, preserving table structure."""

    def parse(self, file_path: str) -> list[Document]:
        documents = []
        path = Path(file_path)

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text_parts = []

                tables = page.extract_tables()
                table_bboxes = []
                for table in tables:
                    if table:
                        header = table[0]
                        rows = table[1:]
                        col_widths = [max(len(str(cell or "")) for cell in col) for col in zip(*table)]
                        table_text = " | ".join(str(h or "") for h in header) + "\n"
                        table_text += "-" * sum(col_widths) + "\n"
                        for row in rows:
                            table_text += " | ".join(str(cell or "") for cell in row) + "\n"
                        text_parts.append(table_text)

                page_text = page.extract_text() or ""
                if page_text.strip():
                    text_parts.insert(0, page_text)

                full_text = "\n\n".join(text_parts)
                if full_text.strip():
                    documents.append(Document(
                        content=full_text.strip(),
                        metadata={
                            "page": page_num,
                            "total_pages": len(pdf.pages),
                            "filename": path.name,
                            "has_tables": len(tables) > 0,
                        },
                        source=str(path),
                        doc_type="pdf",
                    ))

        return documents


class TXTParser:
    """Parses plain text files, preserving section structure."""

    def parse(self, file_path: str) -> list[Document]:
        path = Path(file_path)
        text = path.read_text(encoding="utf-8")

        sections = self._split_sections(text)
        documents = []
        for i, (title, content) in enumerate(sections):
            documents.append(Document(
                content=content.strip(),
                metadata={
                    "section_index": i,
                    "section_title": title,
                    "filename": path.name,
                    "total_sections": len(sections),
                },
                source=str(path),
                doc_type="txt",
            ))
        return documents

    def _split_sections(self, text: str) -> list[tuple[str, str]]:
        lines = text.split("\n")
        sections = []
        current_title = "Header"
        current_content = []

        for line in lines:
            stripped = line.strip()
            if stripped.startswith("===") and stripped.endswith("==="):
                if current_content:
                    sections.append((current_title, "\n".join(current_content)))
                current_title = stripped.strip("= ")
                current_content = []
            else:
                current_content.append(line)

        if current_content:
            sections.append((current_title, "\n".join(current_content)))

        return sections


class ExcelParser:
    """Parses Excel files, converting each sheet into structured text."""

    def parse(self, file_path: str) -> list[Document]:
        path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        documents = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h or "") for h in rows[0]]
            text_parts = [f"Sheet: {sheet_name}", f"Columns: {', '.join(headers)}", ""]

            for row in rows[1:]:
                row_text = " | ".join(
                    f"{headers[i]}: {self._format_cell(cell)}"
                    for i, cell in enumerate(row) if i < len(headers)
                )
                text_parts.append(row_text)

            summary = self._generate_sheet_summary(headers, rows[1:])
            text_parts.insert(3, f"\nSummary: {summary}\n")

            documents.append(Document(
                content="\n".join(text_parts),
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(rows) - 1,
                    "columns": headers,
                    "filename": path.name,
                },
                source=str(path),
                doc_type="excel",
            ))

        wb.close()
        return documents

    def _format_cell(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return f"{value:.4f}" if abs(value) < 1 else f"{value:.2f}"
        return str(value)

    def _generate_sheet_summary(self, headers: list[str], rows: list) -> str:
        numeric_cols = {}
        for i, h in enumerate(headers):
            vals = []
            for row in rows:
                if i < len(row) and isinstance(row[i], (int, float)):
                    vals.append(row[i])
            if vals:
                numeric_cols[h] = {
                    "min": min(vals), "max": max(vals),
                    "avg": sum(vals) / len(vals), "count": len(vals),
                }

        parts = [f"{len(rows)} records"]
        for col, stats in list(numeric_cols.items())[:3]:
            parts.append(f"{col}: avg={stats['avg']:.1f}, range=[{stats['min']}, {stats['max']}]")
        return "; ".join(parts)


class DocumentIngestion:
    """Unified document ingestion that auto-detects file type and routes to appropriate parser."""

    def __init__(self):
        self.parsers = {
            ".pdf": PDFParser(),
            ".txt": TXTParser(),
            ".xlsx": ExcelParser(),
            ".xls": ExcelParser(),
        }

    def ingest_file(self, file_path: str) -> list[Document]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = path.suffix.lower()
        parser = self.parsers.get(ext)
        if not parser:
            raise ValueError(f"Unsupported file type: {ext}. Supported: {list(self.parsers.keys())}")

        docs = parser.parse(file_path)
        print(f"  Ingested {path.name}: {len(docs)} document segments")
        return docs

    def ingest_directory(self, dir_path: str) -> list[Document]:
        all_docs = []
        dir_p = Path(dir_path)
        supported = set(self.parsers.keys())

        for file_path in sorted(dir_p.iterdir()):
            if file_path.suffix.lower() in supported:
                all_docs.extend(self.ingest_file(str(file_path)))

        print(f"\nTotal: {len(all_docs)} document segments from {dir_p}")
        return all_docs
