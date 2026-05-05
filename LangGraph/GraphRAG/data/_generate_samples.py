"""Regenerate the sample Excel and PDF manuals under ``data/``.

Run from the repo root (or any cwd):

    python LangGraph/GraphRAG/data/_generate_samples.py

The JSON manual (``sample_manual.json``) is hand-curated and not regenerated
here. The two outputs below are deterministic so re-running the script
produces byte-identical Excel content; the PDF will differ slightly because
ReportLab embeds a creation timestamp.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors

DATA_DIR = Path(__file__).resolve().parent


# ---------------------------------------------------------------------------
# Excel: pump spare parts catalog
# ---------------------------------------------------------------------------
PARTS_ROWS = [
    # (part_id, name, category, compatible_model, lead_time_days, unit_price_usd, notes)
    ("P-1001", "Mechanical seal", "Sealing",      "CP-200 / CP-250", 14,  185.00,
     "Replace if leakage exceeds 5 ml/hr; PART_OF annual seal replacement."),
    ("P-1002", "Deep groove ball bearing", "Bearing", "CP-200 series",  7,   62.50,
     "REQUIRES re-greasing every 1000 hours; replace at first sign of vibration."),
    ("P-1003", "Impeller (cast iron)", "Hydraulic", "CP-250",           21,  410.00,
     "Inspect for cavitation pitting; CAUSED_BY low NPSHa."),
    ("P-1004", "Foot valve, 2 inch", "Suction",    "All centrifugals",  5,   78.00,
     "Clogged foot valve is a common SYMPTOM_OF loss of prime."),
    ("P-1005", "Thermal overload relay", "Electrical", "Motor starter MS-30", 3, 92.00,
     "Tripped relay is a SYMPTOM_OF seized impeller or excessive load."),
    ("P-1006", "Coupling alignment shim kit", "Alignment", "Universal",  2,   24.00,
     "Used when MITIGATING excessive vibration via re-alignment."),
    ("P-1007", "Cooling fan, axial 120mm", "Cooling", "Motor frame 132M", 5, 45.00,
     "Blocked fan CAUSED_BY dust buildup leads to motor overheating."),
    ("P-1008", "O-ring set (Viton)", "Sealing",    "CP-200 / CP-250",   2,   12.50,
     "Replace at every overhaul; PART_OF preventive maintenance plan."),
    ("P-1009", "Pressure gauge, 0-10 bar", "Instrumentation", "Universal", 4, 38.00,
     "Calibrate annually; RELATED_TO discharge head verification."),
    ("P-1010", "Suction strainer, SS304", "Suction", "All centrifugals", 6, 56.00,
     "Prevents debris ingress; reduces risk of impeller damage."),
]

PARTS_COLUMNS = [
    "part_id", "name", "category", "compatible_model",
    "lead_time_days", "unit_price_usd", "notes",
]


def write_parts_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Spare Parts"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(PARTS_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in PARTS_ROWS:
        ws.append(list(row))

    widths = [10, 28, 16, 22, 16, 16, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Second sheet: vendor contacts (so the file has multi-sheet content)
    vendors = wb.create_sheet("Vendors")
    vendors.append(["vendor_id", "name", "region", "lead_time_days", "specialty"])
    vendors.append(["V-01", "AcmePump Spares",  "APAC",  10, "Hydraulics, impellers"])
    vendors.append(["V-02", "ElectroCore Ltd.", "EMEA",  6,  "Motor starters, relays"])
    vendors.append(["V-03", "SealTech Inc.",    "AMER",  8,  "Mechanical seals, O-rings"])
    for cell in vendors[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    wb.save(path)


# ---------------------------------------------------------------------------
# PDF: electrical safety / lockout-tagout manual
# ---------------------------------------------------------------------------
SAFETY_SECTIONS = [
    (
        "1. Purpose and scope",
        "This manual defines the lockout/tagout (LOTO) procedure REQUIRED before any "
        "maintenance or inspection task on industrial pumps and their driving motors. "
        "It applies to all certified technicians and PRECEDES any task documented in "
        "the Industrial Pump Troubleshooting Manual.",
    ),
    (
        "2. Hazards addressed",
        "Unexpected energization of a motor is CAUSED_BY residual capacitance, gravity, "
        "stored fluid pressure, or improper isolation. These hazards are RELATED_TO the "
        "majority of recordable injuries in pump maintenance and can be a SYMPTOM_OF "
        "skipping the LOTO sequence below.",
    ),
    (
        "3. Required personal protective equipment",
        "PPE for LOTO REQUIRES insulated gloves rated for the system voltage, arc-flash "
        "face shield, dielectric footwear, and a personal padlock with a unique key. "
        "PPE is PART_OF the daily pre-task checklist and must be inspected before use.",
    ),
    (
        "4. Lockout/Tagout procedure",
        "Step 1 — Notify all affected personnel. "
        "Step 2 — Identify all energy sources (electrical, hydraulic, pneumatic). "
        "Step 3 — Shut down the equipment using normal stopping procedure. "
        "Step 4 — Isolate each energy source at its disconnect. "
        "Step 5 — Apply your personal lock and tag at every isolation point. "
        "Step 6 — Release stored energy (bleed pressure, discharge capacitors). "
        "Step 7 — Verify zero energy with a calibrated meter before starting work. "
        "Skipping any step is CAUSED_BY time pressure and is the top contributor to "
        "near-miss incidents.",
    ),
    (
        "5. Restoring service",
        "After maintenance, restoration PRECEDES the next production run and REQUIRES "
        "the following: clear the work area, reinstall guards, remove personal locks "
        "(only the owner may remove their own lock), notify operators, and re-energize "
        "the system at the disconnect. Verify normal operation before leaving the area.",
    ),
    (
        "6. Incident response",
        "If an unexpected release of energy occurs, evacuate the zone, isolate the source "
        "from a safe distance, and notify the shift supervisor. Document the event within "
        "24 hours; root-cause analysis is PART_OF the continuous improvement program and "
        "is RELATED_TO updates of this manual.",
    ),
]

EMERGENCY_CONTACTS = [
    ["Role", "Name", "Phone", "Availability"],
    ["Shift Supervisor",       "On duty",        "Ext. 101", "24x7"],
    ["EHS Officer",            "P. Ramanathan",  "Ext. 215", "Mon-Fri 08:00-18:00"],
    ["Plant Electrician",      "On call",        "Ext. 330", "24x7"],
    ["External Emergency",     "Local fire dept", "112",     "24x7"],
]


def write_safety_pdf(path: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=18,
    )
    h2 = ParagraphStyle(
        "H2",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#1F4E79"),
    )
    body = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontSize=11,
        leading=15,
        spaceAfter=6,
    )

    doc = SimpleDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.9 * inch,
        rightMargin=0.9 * inch,
        topMargin=0.8 * inch,
        bottomMargin=0.8 * inch,
        title="Electrical Safety & Lockout-Tagout Manual",
        author="GraphRAG Sample Corpus",
    )

    story = [
        Paragraph("Electrical Safety &amp; Lockout-Tagout Manual", title_style),
        Paragraph("Document SAF-LOTO-01 &middot; Revision 1.0 &middot; Sample corpus for GraphRAG", subtitle_style),
    ]

    for heading, text in SAFETY_SECTIONS:
        story.append(Paragraph(heading, h2))
        story.append(Paragraph(text, body))

    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("7. Emergency contacts", h2))
    table = Table(EMERGENCY_CONTACTS, colWidths=[1.6 * inch, 1.7 * inch, 1.2 * inch, 1.8 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
                ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
                ("INNERGRID",  (0, 0), (-1, -1), 0.25, colors.grey),
                ("BOX",        (0, 0), (-1, -1), 0.5,  colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("FONTSIZE",   (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
            ]
        )
    )
    story.append(table)

    doc.build(story)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = DATA_DIR / "sample_parts_catalog.xlsx"
    pdf_path = DATA_DIR / "sample_safety_manual.pdf"
    write_parts_xlsx(xlsx_path)
    write_safety_pdf(pdf_path)
    print(f"wrote {xlsx_path}")
    print(f"wrote {pdf_path}")


if __name__ == "__main__":
    main()
